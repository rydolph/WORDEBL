import openpyxl
import re

# Укажите путь к исходному файлу и выходному файлу
input_file = 'output.xlsx'  # Замените на ваш файл
output_file = 'your_output_file.xlsx'  # Замените на имя выходного файла

# Функция для обработки текста с диапазонами
def process_range(cell_value):
    try:
        # Ищем все диапазоны в формате "число - число"
        pattern = r'(\d{1,2})\s*-\s*(\d{1,2})'
        matches = re.finditer(pattern, cell_value)

        # Заменяем диапазоны на перечисления
        for match in matches:
            start, end = map(int, match.groups())
            replacement = ', '.join(map(lambda x: f'{x:02}', range(start, end + 1)))
            cell_value = cell_value.replace(match.group(0), replacement)

        return cell_value
    except Exception as e:
        print(f"Ошибка при обработке строки: {cell_value}. Ошибка: {e}")
        return cell_value

# Открываем Excel-файл
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# Обрабатываем данные из столбца B и записываем в столбец D
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
    cell = row[0]
    if cell.value:  # Проверяем, что значение не пустое
        processed_value = process_range(cell.value.strip())  # Убираем пробелы вокруг значения
        sheet.cell(row=cell.row, column=4, value=processed_value)  # Записываем результат в столбец D

# Сохраняем изменения в новый файл
workbook.save(output_file)
print(f"Файл успешно обработан и сохранён как {output_file}")
