import openpyxl

def process_data(input_file, output_file):
    # Открываем Excel файл
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Получаем максимальное количество строк изначально
    max_row = sheet.max_row

    # Проходим по всем строкам в столбце F
    row = 2  # Предполагаем, что первая строка — заголовок
    while row <= max_row:
        value = sheet[f"F{row}"].value
        if value and "," in value:  # Если есть перечисление
            numbers = [num.strip() for num in value.split(",")]
            for i, num in enumerate(numbers):
                if i == 0:
                    # Первое значение остается на месте
                    sheet[f"G{row}"].value = num
                else:
                    # Создаем новую строку для остальных значений
                    max_row += 1
                    sheet.insert_rows(row + 1)
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row + 1, column=col).value = sheet.cell(row=row, column=col).value
                    sheet[f"G{row + 1}"].value = num
                    row += 1
        else:
            # Если перечисления нет, просто переносим значение в G
            sheet[f"G{row}"].value = value
        row += 1

    # Сохраняем изменения в новый файл
    wb.save(output_file)

# Пример использования
input_file = "kogotbobra.xlsx"  # Имя входного файла
output_file = "pizdec.xlsx"  # Имя выходного файла
process_data(input_file, output_file)
